"""
Social-engineering / policy-violation detection pipeline (stage 2).

Reads conversations produced by the generation pipeline and runs each turn
through one of two detectors based on speaker:

  - Caller turns        -> social-engineering detector
  - Representative turns (excluding the turn-0 greeting) -> policy-violation detector

For each detected turn, the chosen detector is run three times — once per
"stance" (high_precision / high_recall / balanced) — yielding 6 (objective,
stance) combinations stored as parallel per-turn lists per conversation.

Output behavior:
  - One row per conversation in the detection metadata xlsx, mirroring the
    generation metadata grain.
  - Generation-phase metadata columns are preserved with `generation_`
    prefix on those whose meaning is ambiguous between phases (model,
    temperature, status, attempts, tokens, etc.).
  - Per-turn detection data is stored as JSON-serialized lists in the
    metadata xlsx and as nested objects in the aggregated detections JSON.

API call shape:
  max_tokens=1, logprobs=True, top_logprobs=10. Temperature and top_p are
  left at API defaults — with logprobs we get full distributional
  information regardless of temperature, so determinism is irrelevant.
  No logit_bias — we want unbiased base confidence levels. The top 10
  tokens with their logprobs are scanned in descending order; the first
  token that is exactly "Y" or "N" is taken as the prediction. If no Y
  or N appears in the top 10, prediction is recorded as "E" (model output
  outside the expected vocabulary, likely a prompt or model issue).

  Note: this is set up for the GPT-4.1 family (which uses the legacy
  `max_tokens` parameter and supports logprobs). The GPT-5 family
  deprecated logprobs and uses `max_completion_tokens` instead, so it
  would require a different call shape if used.

Per-turn prediction codes:
  "Y" -> detected
  "N" -> not detected
  "E" -> model produced no Y/N in top 10 (no retry — this is diagnostic)
  "X" -> call failed entirely after retries (placeholder so list lengths
         remain aligned across all per-turn fields)
  None -> turn not applicable to this detector (turn 0, or wrong speaker)

Retries are minimal: total of 2 attempts (one initial + one retry) for
genuine API errors. "E" outputs are NOT retried.
"""

from __future__ import annotations

import json
import math
import os
import shutil
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from tqdm.auto import tqdm

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

STANCES = ("high_precision", "high_recall", "balanced")
OBJECTIVES = ("policy_violation", "social_engineering")

# (objective, stance) keys used as column / json keys for parallel per-turn lists
DETECTION_KEYS = [f"{obj}__{st}" for obj in OBJECTIVES for st in STANCES]

# Per-turn fields recorded for each detection key. ALL lists have length
# n_turns and are filled positionally. Non-applicable turns get None.
PER_TURN_FIELDS = (
    "prediction",      # "Y" | "N" | "E" | "X" | None (None = not applicable)
    "p_detected",      # exp(logprob) for "Y", 3 decimals; None if not applicable
    "p_not_detected",  # exp(logprob) for "N", 3 decimals; None if not applicable
    "latency_ms",      # int ms, or None if call failed / not applicable
    "input_tokens",    # int, or None
    "output_tokens",   # int, or None
    "top10",           # [[token, prob], ...] up to 10 entries; None if not applicable
)

DEFAULT_MAX_ATTEMPTS = 2     # one initial + one retry, for genuine API errors only
DEFAULT_BACKOFF = (4,)        # single retry waits this long
DEFAULT_CONCURRENCY = 4
DEFAULT_TOP_LOGPROBS = 10

# Prediction code constants
PRED_DETECTED = "Y"
PRED_NOT_DETECTED = "N"
PRED_NO_VALID = "E"   # no Y or N in top-10
PRED_FAILED = "X"     # API call failed after retries

# Generation-phase columns that get a `generation_` prefix when carried over
GENERATION_AMBIGUOUS_COLUMNS = {
    "model", "temperature", "top_p", "status", "attempts",
    "last_error", "generated_at_utc",
    "input_tokens", "output_tokens", "total_tokens",
}

# Generation-phase columns that are kept as-is (they unambiguously describe
# the *generation request*, not the detection request).
GENERATION_PASSTHROUGH_COLUMNS = {
    "request_id", "replicate_index", "prompt_template_key",
    "scenario", "representative", "caller", "benign_context",
    "cialdini_emphasis", "turn_count_value", "flavor",
    "system_prompt", "user_prompt",
}


# ---------------------------------------------------------------------------
# Loading dimension files and assembling detector prompts
# ---------------------------------------------------------------------------

def load_detection_dimensions(path: str | Path) -> dict:
    """Load a detection prompt dimensions JSON file."""
    with open(path, "r", encoding="utf-8") as f:
        d = json.load(f)
    # Light validation
    for k in ("objective_key", "applies_to_speaker", "base_prompt", "stances"):
        if k not in d:
            raise ValueError(f"Missing key {k!r} in {path}")
    if "template" not in d["base_prompt"]:
        raise ValueError(f"base_prompt.template missing in {path}")
    for st in STANCES:
        if st not in d["stances"]:
            raise ValueError(f"Missing stance {st!r} in {path}")
        if "instruction" not in d["stances"][st]:
            raise ValueError(f"stances.{st}.instruction missing in {path}")
    return d


def assemble_system_prompt(dims: dict, stance: str) -> str:
    """Return the system prompt for one (objective, stance) combination."""
    template = dims["base_prompt"]["template"]
    instruction = dims["stances"][stance]["instruction"]
    return template.replace("{stance_instruction}", instruction)


# ---------------------------------------------------------------------------
# Conversation IO
# ---------------------------------------------------------------------------

def load_conversations(path: str | Path) -> dict:
    """Load the aggregated generation-phase conversations JSON."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def speaker_for_turn(i: int) -> str:
    """
    The generation pipeline encodes speaker by index parity:
      turn 0  -> representative (greeting)
      odd     -> caller
      even>=2 -> representative
    """
    if i == 0:
        return "representative"
    return "caller" if (i % 2 == 1) else "representative"


def ordered_turn_keys(conversation_dict: dict) -> list[str]:
    """
    Return the conversation's turn keys (e.g. ['turn_0','turn_1',...]) in
    numeric order. Defends against unexpected key ordering in the JSON.
    """
    keys = list(conversation_dict.keys())
    # Each key is "turn_<N>". Sort by the integer suffix.
    def idx(k: str) -> int:
        try:
            return int(k.split("_", 1)[1])
        except (IndexError, ValueError):
            return -1
    keys.sort(key=idx)
    # Sanity: keys should be turn_0, turn_1, ... contiguous from 0
    for expected, k in enumerate(keys):
        if k != f"turn_{expected}":
            raise ValueError(
                f"unexpected turn key sequence: got {keys!r}, "
                f"expected turn_0..turn_{len(keys)-1}"
            )
    return keys


def transcript_through_turn(
    conversation_dict: dict,
    turn_keys: list[str],
    target_turn_index: int,
) -> str:
    """
    Format the conversation through `target_turn_index` (inclusive) as a
    plain transcript with ONLY speaker + passage. No labels, no metadata.

    `conversation_dict` is the generation-format dict shaped like
    {"turn_0": {"passage": "..."}, "turn_1": {"passage": "...", ...}, ...}.
    """
    lines = []
    for i in range(target_turn_index + 1):
        turn = conversation_dict[turn_keys[i]]
        passage = turn["passage"]
        label = "Representative" if speaker_for_turn(i) == "representative" else "Caller"
        lines.append(f"{label}: {passage}")
    return "\n\n".join(lines)


# ---------------------------------------------------------------------------
# OpenAI call
# ---------------------------------------------------------------------------

def _logprob_to_prob(lp: float) -> float:
    """Convert a logprob to a probability, rounded to 3 decimals."""
    return round(math.exp(lp), 3)


def _extract_top10(top_logprobs_list: list) -> list[list]:
    """
    Convert the SDK's top_logprobs list into [[token, prob], ...] form,
    sorted descending by probability, with probabilities rounded to 3 dp.
    """
    pairs = [(alt.token, alt.logprob) for alt in top_logprobs_list]
    # Already sorted by the API, but sort defensively
    pairs.sort(key=lambda x: x[1], reverse=True)
    return [[tok, _logprob_to_prob(lp)] for tok, lp in pairs]


def call_detector(
    client: OpenAI,
    system_prompt: str,
    user_prompt: str,
    model: str,
    top_logprobs: int = DEFAULT_TOP_LOGPROBS,
) -> dict:
    """
    Run one detection call. No logit_bias is applied — this measures the
    model's unbiased confidence between Y, N, and any other tokens.

    Returns a dict with:
      prediction:     "Y" | "N" | "E"
      p_detected:     probability of "Y" in top-10 (0.0 if absent), 3 dp
      p_not_detected: probability of "N" in top-10 (0.0 if absent), 3 dp
      latency_ms:     int wall-clock around the API call
      input_tokens, output_tokens: ints from response usage
      top10:          list of [token, prob] pairs, descending by prob

    Prediction selection:
      Walk the top-10 in descending logprob order; the first token that is
      exactly "Y" or "N" wins. If neither appears in the top-10, prediction
      is "E" (and p_detected / p_not_detected are 0.0 / 0.0).

    This function does NOT retry on "E" — that's a model/prompt diagnostic,
    not a transient failure.
    """
    t0 = time.perf_counter()
    resp = client.chat.completions.create(
        model=model,
        # The GPT-4.1 family uses the legacy `max_tokens` parameter
        # (the GPT-5 / o-series deprecated it for `max_completion_tokens`).
        # Temperature and top_p are left at defaults — with logprobs we have
        # full distributional information regardless of temperature.
        max_tokens=1,
        logprobs=True,
        top_logprobs=top_logprobs,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    )
    latency_ms = int(round((time.perf_counter() - t0) * 1000))

    choice = resp.choices[0]
    if choice.logprobs is None or not choice.logprobs.content:
        raise RuntimeError("Detector response missing logprobs")

    lp_entry = choice.logprobs.content[0]
    top10 = _extract_top10(lp_entry.top_logprobs)

    # Build a token -> prob lookup for quick access
    token_to_prob = {tok: prob for tok, prob in top10}

    # Walk top-10 in order; first exact "Y" or "N" wins
    prediction = PRED_NO_VALID
    for tok, _prob in top10:
        if tok == "Y":
            prediction = PRED_DETECTED
            break
        if tok == "N":
            prediction = PRED_NOT_DETECTED
            break

    p_y = token_to_prob.get("Y", 0.0)
    p_n = token_to_prob.get("N", 0.0)

    usage = resp.usage
    input_tokens = getattr(usage, "prompt_tokens", 0) if usage else 0
    output_tokens = getattr(usage, "completion_tokens", 0) if usage else 0

    return {
        "prediction": prediction,
        "p_detected": p_y,
        "p_not_detected": p_n,
        "latency_ms": latency_ms,
        "input_tokens": int(input_tokens),
        "output_tokens": int(output_tokens),
        "top10": top10,
    }


def call_detector_with_retries(
    client: OpenAI,
    system_prompt: str,
    user_prompt: str,
    model: str,
    max_attempts: int = DEFAULT_MAX_ATTEMPTS,
    backoff: tuple[int, ...] = DEFAULT_BACKOFF,
    top_logprobs: int = DEFAULT_TOP_LOGPROBS,
) -> tuple[dict | None, int, str]:
    """
    Wrap call_detector with retries on genuine API errors.

    "E" predictions (no Y/N in top-10) are NOT retried — they are returned
    immediately as successful calls. Only exceptions trigger retry.

    Returns (result_dict_or_None, attempts_used, last_error_message).
    On a complete failure (all attempts raised), result is None.
    """
    last_err = ""
    for attempt in range(1, max_attempts + 1):
        try:
            res = call_detector(
                client=client,
                system_prompt=system_prompt,
                user_prompt=user_prompt,
                model=model,
                top_logprobs=top_logprobs,
            )
            return res, attempt, ""
        except Exception as e:
            last_err = f"{type(e).__name__}: {e}"
            if attempt >= max_attempts:
                break
            sleep_for = backoff[min(attempt - 1, len(backoff) - 1)]
            time.sleep(sleep_for)
    return None, max_attempts, last_err


# ---------------------------------------------------------------------------
# Per-conversation orchestration
# ---------------------------------------------------------------------------

def empty_per_turn_lists(n_turns: int) -> dict:
    """Initialize one (objective, stance)'s per-turn fields with all None."""
    return {field: [None] * n_turns for field in PER_TURN_FIELDS}


def detect_conversation(
    client: OpenAI,
    conversation_id: str,
    conversation_record: dict,
    policy_dims: dict,
    se_dims: dict,
    model: str,
    max_attempts: int = DEFAULT_MAX_ATTEMPTS,
    backoff: tuple[int, ...] = DEFAULT_BACKOFF,
    top_logprobs: int = DEFAULT_TOP_LOGPROBS,
) -> dict:
    """
    Run all applicable detections for a single conversation.

    Returns a dict shaped like:
      {
        "conversation_id": <id>,
        "n_turns": int,
        "speakers": [...],
        "detections": { "<obj>__<stance>": { <field>: [per-turn list], ... } },
        "errors": [ {"target_turn": i, "key": "...", "message": "..."} , ... ],
      }

    Per-turn list lengths are guaranteed equal to n_turns and aligned across
    all fields and all detection keys. Non-applicable turns get None for
    every field. Failed calls get "X" for prediction and None for the rest.
    """
    # The generation pipeline stores `conversation` as a dict shaped like
    # {"turn_0": {"passage": ...}, "turn_1": {"passage": ..., ...}, ...}.
    # Speaker is encoded by index parity, not stored in the turn itself.
    conv_dict = conversation_record["conversation"]
    turn_keys = ordered_turn_keys(conv_dict)
    n_turns = len(turn_keys)
    speakers = [speaker_for_turn(i) for i in range(n_turns)]

    # Pre-build system prompts for all 6 keys
    system_prompts = {}
    for stance in STANCES:
        system_prompts[f"policy_violation__{stance}"] = assemble_system_prompt(policy_dims, stance)
        system_prompts[f"social_engineering__{stance}"] = assemble_system_prompt(se_dims, stance)

    detections = {key: empty_per_turn_lists(n_turns) for key in DETECTION_KEYS}
    errors = []

    # Walk each turn; pick the right objective; run all 3 stances.
    for i in range(n_turns):
        if i == 0:
            # Turn 0 is the representative greeting; skip both detectors.
            continue

        speaker = speakers[i]
        if speaker == "representative":
            objective = "policy_violation"
        elif speaker == "caller":
            objective = "social_engineering"
        else:
            errors.append({"target_turn": i, "key": None,
                           "message": f"unknown speaker {speaker!r}"})
            continue

        user_prompt = transcript_through_turn(conv_dict, turn_keys, i)

        for stance in STANCES:
            key = f"{objective}__{stance}"
            res, _attempts, err = call_detector_with_retries(
                client=client,
                system_prompt=system_prompts[key],
                user_prompt=user_prompt,
                model=model,
                max_attempts=max_attempts,
                backoff=backoff,
                top_logprobs=top_logprobs,
            )
            if res is None:
                # Genuine call failure after retries. Record "X" for
                # prediction so per-turn list lengths stay aligned; leave
                # the other fields None and surface the error.
                detections[key]["prediction"][i] = PRED_FAILED
                errors.append({"target_turn": i, "key": key, "message": err})
                continue
            # Successful call (which includes "E" — model gave no Y/N in top-10).
            detections[key]["prediction"][i] = res["prediction"]
            detections[key]["p_detected"][i] = res["p_detected"]
            detections[key]["p_not_detected"][i] = res["p_not_detected"]
            detections[key]["latency_ms"][i] = res["latency_ms"]
            detections[key]["input_tokens"][i] = res["input_tokens"]
            detections[key]["output_tokens"][i] = res["output_tokens"]
            detections[key]["top10"][i] = res["top10"]

    # Alignment guarantee: every per-turn list has exactly n_turns entries.
    for key in DETECTION_KEYS:
        for field in PER_TURN_FIELDS:
            assert len(detections[key][field]) == n_turns, (
                f"alignment violation: detections[{key!r}][{field!r}] has "
                f"length {len(detections[key][field])}, expected {n_turns}"
            )

    return {
        "conversation_id": conversation_id,
        "n_turns": n_turns,
        "speakers": speakers,
        "detections": detections,
        "errors": errors,
    }


# ---------------------------------------------------------------------------
# Metadata xlsx — schema and IO
# ---------------------------------------------------------------------------

def detection_only_columns() -> list[str]:
    """
    Detection-specific columns appended to each metadata row.

    Per-turn list columns come first (one per (objective, stance) and
    per_turn field), followed by aggregate detection-run columns.
    """
    cols = []
    # Per-(key, field) JSON-list columns
    for key in DETECTION_KEYS:
        for field in PER_TURN_FIELDS:
            cols.append(f"{key}__{field}")
    # Aggregate / run-level columns
    cols.extend([
        "detection_model",
        "detection_status",         # success | partial | error
        "detection_error_count",    # count of "X" outcomes (call failed after retries)
        "detection_no_valid_count", # count of "E" outcomes (no Y/N in top-10)
        "detection_last_error",
        "detection_started_at_utc",
        "detection_finished_at_utc",
        "detection_total_latency_ms",
        "detection_total_input_tokens",
        "detection_total_output_tokens",
    ])
    return cols


def rename_generation_columns(generation_columns: list[str]) -> list[str]:
    """
    Apply `generation_` prefix to ambiguous columns; pass through the rest.
    Order is preserved.
    """
    out = []
    for c in generation_columns:
        if c in GENERATION_AMBIGUOUS_COLUMNS:
            out.append(f"generation_{c}")
        else:
            out.append(c)
    return out


def build_detection_columns(generation_columns: list[str]) -> list[str]:
    """Renamed generation columns + detection-specific columns."""
    return rename_generation_columns(generation_columns) + detection_only_columns()


def read_generation_metadata(path: str | Path) -> tuple[list[str], list[dict]]:
    """Read generation metadata xlsx. Returns (header_columns, rows)."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = []
    for r in rows_iter:
        if r is None or all(v is None for v in r):
            continue
        rows.append({header[i]: r[i] for i in range(len(header))})
    wb.close()
    return header, rows


def write_detection_metadata_xlsx(
    out_path: str | Path,
    columns: list[str],
    rows: list[dict],
) -> None:
    """Atomic write of the detection metadata xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "detections"

    # Header
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top")

    # Body
    for i, row in enumerate(rows, start=2):
        for j, col in enumerate(columns, start=1):
            v = row.get(col, "")
            if isinstance(v, (list, dict)):
                v = json.dumps(v, separators=(",", ":"))
            ws.cell(row=i, column=j, value=v)

    # Reasonable widths
    widths = {}
    for col in columns:
        if col.endswith("__top10"):
            widths[col] = 80
        elif any(col.endswith(suf) for suf in PER_TURN_FIELDS):
            widths[col] = 36
        elif col in GENERATION_PASSTHROUGH_COLUMNS or col == "request_id":
            widths[col] = 24
        elif col.startswith("detection_") or col.startswith("generation_"):
            widths[col] = 18
        else:
            widths[col] = 18
    for j, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = widths.get(col, 18)
    ws.freeze_panes = "A2"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".tmp_det_meta_", suffix=".xlsx",
                               dir=str(out_path.parent))
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, out_path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
        wb.close()


# ---------------------------------------------------------------------------
# Detections JSON store (atomic write)
# ---------------------------------------------------------------------------

def load_detections_store(path: str | Path) -> dict:
    """Load the aggregated detections JSON, or return an empty store."""
    path = Path(path)
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"conversations": {}}


def save_detections_store(path: str | Path, store: dict) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".tmp_det_", suffix=".json",
                               dir=str(path.parent))
    os.close(fd)
    try:
        # ensure_ascii=False keeps the on-disk file readable for any
        # non-ASCII passages; encoding="utf-8" is required on Windows where
        # the default is cp1252.
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(store, f, indent=2, ensure_ascii=False)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)


# ---------------------------------------------------------------------------
# Run-level orchestration
# ---------------------------------------------------------------------------

def conversation_id_for_row(row: dict) -> str:
    """Generation rows are keyed by request_id; that's our conversation_id."""
    rid = row.get("request_id")
    if not rid:
        raise ValueError("Generation metadata row missing request_id")
    return str(rid)


def aggregate_run_metrics(detection_result: dict) -> dict:
    """Compute totals across all (key, turn) cells for a conversation."""
    total_lat = 0
    total_in = 0
    total_out = 0
    for key in DETECTION_KEYS:
        d = detection_result["detections"][key]
        total_lat += sum(v for v in d["latency_ms"] if v is not None)
        total_in += sum(v for v in d["input_tokens"] if v is not None)
        total_out += sum(v for v in d["output_tokens"] if v is not None)
    return {
        "detection_total_latency_ms": total_lat,
        "detection_total_input_tokens": total_in,
        "detection_total_output_tokens": total_out,
    }


def detection_status_for(detection_result: dict) -> str:
    """
    Classify the run as success / partial / error.

    success: every applicable turn got a Y, N, or E (no X failures).
    partial: at least one applicable turn got a result, but at least one
             also got X (call failed).
    error:   no applicable turns got any result (every applicable slot is X
             — meaning all calls failed).
    """
    n_total_applicable = 0
    n_x = 0
    n_resolved = 0  # Y / N / E
    for key in DETECTION_KEYS:
        for v in detection_result["detections"][key]["prediction"]:
            if v is None:
                continue
            n_total_applicable += 1
            if v == PRED_FAILED:
                n_x += 1
            else:
                n_resolved += 1

    if n_total_applicable == 0:
        # Nothing applicable was attempted (e.g., conversation had only one turn).
        # Treat as success: there was nothing to detect.
        return "success"
    if n_x == 0:
        return "success"
    if n_resolved == 0:
        return "error"
    return "partial"


def build_detection_row(
    generation_row: dict,
    detection_result: dict,
    model: str,
    started_at: str,
    finished_at: str,
) -> dict:
    """Build one full detection-metadata row from a generation row + result."""
    out: dict[str, Any] = {}

    # Carry generation columns with renaming
    for c, v in generation_row.items():
        if c in GENERATION_AMBIGUOUS_COLUMNS:
            out[f"generation_{c}"] = v
        else:
            out[c] = v

    # Per-turn list columns
    for key in DETECTION_KEYS:
        d = detection_result["detections"][key]
        for field in PER_TURN_FIELDS:
            out[f"{key}__{field}"] = d[field]

    # Aggregate columns
    totals = aggregate_run_metrics(detection_result)

    # Count "E" and "X" prediction codes across all (key, turn) slots
    n_e = 0
    n_x = 0
    for key in DETECTION_KEYS:
        for v in detection_result["detections"][key]["prediction"]:
            if v == PRED_NO_VALID:
                n_e += 1
            elif v == PRED_FAILED:
                n_x += 1

    out["detection_model"] = model
    out["detection_status"] = detection_status_for(detection_result)
    out["detection_error_count"] = n_x
    out["detection_no_valid_count"] = n_e
    out["detection_last_error"] = (
        detection_result["errors"][-1]["message"] if detection_result["errors"] else ""
    )
    out["detection_started_at_utc"] = started_at
    out["detection_finished_at_utc"] = finished_at
    out.update(totals)

    return out


def run_detection_pipeline(
    *,
    conversations_path: str | Path,
    generation_metadata_path: str | Path,
    detections_json_path: str | Path,
    detection_metadata_path: str | Path,
    policy_dims_path: str | Path,
    se_dims_path: str | Path,
    model: str,
    client: OpenAI | None = None,
    max_requests: int | None = None,
    concurrency: int = DEFAULT_CONCURRENCY,
    max_attempts: int = DEFAULT_MAX_ATTEMPTS,
    backoff: tuple[int, ...] = DEFAULT_BACKOFF,
    top_logprobs: int = DEFAULT_TOP_LOGPROBS,
    flush_every: int = 10,
    skip_already_done: bool = True,
) -> dict:
    """
    Run detection for every conversation in `conversations_path`.

    Writes/updates `detections_json_path` and `detection_metadata_path`
    incrementally. Returns a small summary dict.

    Set `skip_already_done=True` (default) to resume: any conversation
    that already has a fully-successful row in the existing detection
    metadata will be skipped.
    """
    client = client or OpenAI()

    policy_dims = load_detection_dimensions(policy_dims_path)
    se_dims = load_detection_dimensions(se_dims_path)

    convs_data = load_conversations(conversations_path)
    convs = convs_data.get("conversations", convs_data)

    gen_columns, gen_rows = read_generation_metadata(generation_metadata_path)
    detection_columns = build_detection_columns(gen_columns)

    # Load any existing partial output to support resume
    store = load_detections_store(detections_json_path)
    existing_rows: list[dict] = []
    done_ids: set[str] = set()
    if Path(detection_metadata_path).exists():
        _, existing_rows = read_generation_metadata(detection_metadata_path)
        if skip_already_done:
            done_ids = {
                r["request_id"]
                for r in existing_rows
                if r.get("detection_status") == "success"
            }

    # Plan the work: one detection job per generation row that has a
    # corresponding conversation and isn't already done.
    pending: list[tuple[dict, dict]] = []
    for gr in gen_rows:
        cid = conversation_id_for_row(gr)
        if cid in done_ids:
            continue
        if cid not in convs:
            # No corresponding conversation in the JSON — skip silently;
            # this is normal for in-flight or failed generations.
            continue
        pending.append((gr, convs[cid]))

    if max_requests is not None:
        pending = pending[:max_requests]

    # Existing rows we will keep (by request_id), updated as new ones land
    by_id: dict[str, dict] = {r["request_id"]: r for r in existing_rows}

    summary = {
        "n_planned": len(pending),
        "n_success": 0,
        "n_partial": 0,
        "n_error": 0,
        "n_calls": 0,
        "n_y": 0,
        "n_n": 0,
        "n_e": 0,
        "n_x": 0,
    }

    def worker(gen_row: dict, conv_record: dict) -> tuple[dict, dict]:
        """Run detection on one conversation; return (gen_row, det_result_with_meta)."""
        cid = conversation_id_for_row(gen_row)
        started = datetime.now(timezone.utc).isoformat(timespec="seconds")
        det_result = detect_conversation(
            client=client,
            conversation_id=cid,
            conversation_record=conv_record,
            policy_dims=policy_dims,
            se_dims=se_dims,
            model=model,
            max_attempts=max_attempts,
            backoff=backoff,
            top_logprobs=top_logprobs,
        )
        finished = datetime.now(timezone.utc).isoformat(timespec="seconds")
        return gen_row, {
            "result": det_result,
            "started_at": started,
            "finished_at": finished,
        }

    def flush():
        # Snapshot ordered rows: prefer original generation order
        ordered = []
        seen = set()
        for gr in gen_rows:
            rid = conversation_id_for_row(gr)
            if rid in by_id:
                ordered.append(by_id[rid])
                seen.add(rid)
        # Append any rows not in current generation file (defensive)
        for rid, row in by_id.items():
            if rid not in seen:
                ordered.append(row)
        write_detection_metadata_xlsx(detection_metadata_path,
                                      detection_columns, ordered)
        save_detections_store(detections_json_path, store)

    completed_since_flush = 0

    with ThreadPoolExecutor(max_workers=concurrency) as ex:
        futures = [ex.submit(worker, gr, cr) for gr, cr in pending]
        for fut in tqdm(as_completed(futures), total=len(futures),
                        desc="detecting", unit="conv"):
            gen_row, payload = fut.result()
            res = payload["result"]
            cid = res["conversation_id"]

            row = build_detection_row(
                generation_row=gen_row,
                detection_result=res,
                model=model,
                started_at=payload["started_at"],
                finished_at=payload["finished_at"],
            )
            by_id[cid] = row

            store["conversations"][cid] = {
                "n_turns": res["n_turns"],
                "speakers": res["speakers"],
                "detections": res["detections"],
                "errors": res["errors"],
                "model": model,
                "started_at_utc": payload["started_at"],
                "finished_at_utc": payload["finished_at"],
            }

            status = row["detection_status"]
            if status == "success":
                summary["n_success"] += 1
            elif status == "partial":
                summary["n_partial"] += 1
            else:
                summary["n_error"] += 1

            # Every applicable (turn, stance) slot has a non-None prediction
            # code (Y, N, E, or X). Tally them for the summary.
            for key in DETECTION_KEYS:
                for v in res["detections"][key]["prediction"]:
                    if v is None:
                        continue
                    summary["n_calls"] += 1
                    if v == PRED_DETECTED:
                        summary["n_y"] += 1
                    elif v == PRED_NOT_DETECTED:
                        summary["n_n"] += 1
                    elif v == PRED_NO_VALID:
                        summary["n_e"] += 1
                    elif v == PRED_FAILED:
                        summary["n_x"] += 1

            completed_since_flush += 1
            if completed_since_flush >= flush_every:
                flush()
                completed_since_flush = 0

    flush()
    return summary
