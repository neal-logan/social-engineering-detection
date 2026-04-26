"""
Adapter layer: the ONLY place this app touches the rest of the project.

Wired against the current `conversation_generation` and `detection_pipeline`
modules. If a signature in either of those changes, edit only this file.

Internal representation note
----------------------------
The on-disk schema stores each conversation as a dict
    record["conversation"] = {"turn_0": {"passage": "...", ...},
                              "turn_1": {"passage": "...", "cialdini_*": ...},
                              ...}
with speaker derived by parity (turn 0 = representative greeting,
odd = caller, even >= 2 = representative response).

The web UI is easier to write against a list-of-turn-dicts where each turn
already carries `speaker` and `text`. So the adapter normalises every
loaded/generated record into that flat list shape under the same key
`record["conversation"]["turns"]`. Original `turn_N` keys remain on
`record["conversation"]` so any downstream code that wants the raw shape
can still get it.
"""

from __future__ import annotations

import json
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import config


# ---------------------------------------------------------------------------
# Make project sibling folders importable.
# ---------------------------------------------------------------------------

_PROJECT_ROOT = config.PROJECT_ROOT
for sub in ("conversation_generation", "detection_pipeline"):
    p = _PROJECT_ROOT / sub
    if p.exists() and str(p) not in sys.path:
        sys.path.insert(0, str(p))
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))


# Lazy imports — both modules need a working .env, so deferring helps the UI
# show a clean error in dynamic mode if something is missing rather than
# failing at module import time.
def _cg():
    import conversation_generation as cg  # noqa: PLC0415
    return cg


def _det():
    import detection as det  # noqa: PLC0415
    return det


# ---------------------------------------------------------------------------
# Schema normalisation: turn_N dict <-> list of turns
# ---------------------------------------------------------------------------

# Module-level diagnostic string set by load_conversations() and any other
# loader to record what was tried during a failure. The UI can surface this
# to help debug missing-file issues.
LAST_LOAD_DIAGNOSTIC: str = ""


def _ordered_turn_keys(conv_dict: dict) -> list[str]:
    """Return turn_0, turn_1, ... in numeric order. Validates contiguity."""
    keys = [k for k in conv_dict.keys() if k.startswith("turn_") and k != "turns"]
    try:
        keys_sorted = sorted(keys, key=lambda k: int(k.split("_", 1)[1]))
    except ValueError as e:
        raise ValueError(f"non-integer suffix in turn key: {e}") from e
    expected = [f"turn_{i}" for i in range(len(keys_sorted))]
    if keys_sorted != expected:
        raise ValueError(
            f"turn keys are not contiguous from 0: got {keys_sorted}, "
            f"expected {expected}"
        )
    return keys_sorted


def _speaker_for_turn(i: int) -> str:
    """Turn 0 = rep greeting, odd = caller, even >= 2 = rep response."""
    return "representative" if i % 2 == 0 else "caller"


def _normalise_record(record: dict) -> dict:
    """
    Mutate-and-return a conversation record so that its conversation block
    has a flat `turns` list with each turn dict carrying `speaker` and `text`
    in addition to whatever flags it already had.
    """
    conv = record.get("conversation")
    if not isinstance(conv, dict):
        return record

    if "turns" in conv and isinstance(conv["turns"], list):
        # Already normalised (e.g., dynamic-mode record we just built).
        return record

    keys = _ordered_turn_keys(conv)
    turns: list[dict[str, Any]] = []
    for i, k in enumerate(keys):
        t = dict(conv[k])  # copy so we don't mutate the original turn dict
        t.setdefault("speaker", _speaker_for_turn(i))
        t.setdefault("text", t.get("passage", ""))
        turns.append(t)
    conv["turns"] = turns
    return record


# ---------------------------------------------------------------------------
# 1. Dimensions
# ---------------------------------------------------------------------------

def load_dimensions() -> dict[str, Any]:
    """Return the parsed prompt_dimensions.json."""
    with open(config.DIMENSIONS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# 2. Pre-generated conversations (static mode source)
# ---------------------------------------------------------------------------

def load_conversations() -> list[dict[str, Any]]:
    """
    Return the full list of conversation records, drawn from any of:
      - threat_conversations.json
      - benign_conversations.json
    in the conversations/ folder. The CONVERSATIONS_PATH config setting can
    point either at a single one of those files or at the conversations/
    directory itself; both are handled here.

    Each record is normalised so the UI sees a flat list of turn dicts under
    record["conversation"]["turns"].

    On failure to find any files, the global LAST_LOAD_DIAGNOSTIC is set with
    a description of what was tried so the UI can surface it.
    """
    global LAST_LOAD_DIAGNOSTIC
    LAST_LOAD_DIAGNOSTIC = ""

    candidates: list[Path] = []
    tried: list[str] = []
    p = config.CONVERSATIONS_PATH
    tried.append(f"configured path: {p}")

    def _scan_dir(d: Path) -> list[Path]:
        found = []
        for name in ("threat_conversations.json", "benign_conversations.json"):
            f = d / name
            tried.append(f"checked: {f} -> {'found' if f.exists() else 'missing'}")
            if f.exists():
                found.append(f)
        # Also accept any other *_conversations.json files in the same dir.
        for f in sorted(d.glob("*_conversations.json")):
            if f not in found:
                tried.append(f"also found: {f}")
                found.append(f)
        return found

    if p.is_dir():
        candidates.extend(_scan_dir(p))
    elif p.exists():
        candidates.append(p)
        if p.name == "threat_conversations.json":
            sib = p.with_name("benign_conversations.json")
            tried.append(f"sibling: {sib} -> {'found' if sib.exists() else 'missing'}")
            if sib.exists():
                candidates.append(sib)
        elif p.name == "benign_conversations.json":
            sib = p.with_name("threat_conversations.json")
            tried.append(f"sibling: {sib} -> {'found' if sib.exists() else 'missing'}")
            if sib.exists():
                candidates.append(sib)
    else:
        # Configured path doesn't exist as file or dir; try its parent.
        if p.parent.is_dir():
            tried.append(f"path missing; falling back to parent dir: {p.parent}")
            candidates.extend(_scan_dir(p.parent))
        else:
            tried.append(f"parent dir also missing: {p.parent}")

    if not candidates:
        LAST_LOAD_DIAGNOSTIC = "\n".join(tried)
        return []

    records: list[dict[str, Any]] = []
    for path in candidates:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        # Generation pipeline writes {"conversations": {request_id: record, ...}}.
        if isinstance(raw, dict) and "conversations" in raw:
            store = raw["conversations"]
        elif isinstance(raw, dict):
            store = raw
        else:
            store = {str(i): r for i, r in enumerate(raw)}
        for rec in store.values():
            rec = _normalise_record(rec)
            label = "threat" if "threat" in path.name else (
                "benign" if "benign" in path.name else path.stem
            )
            rec.setdefault("_source_label", label)
            records.append(rec)

    records.sort(key=lambda r: r.get("generated_at_utc", ""))

    # Filter out conversations whose detection wasn't completed cleanly.
    # Read detection_status from each detection xlsx; drop any record whose
    # id appears in the detection metadata with status != "success".
    # Records that don't appear in any detection xlsx are kept (they haven't
    # been detected against yet — that's not a "broken" state).
    bad_ids = _load_broken_conversation_ids()
    if bad_ids:
        before = len(records)
        records = [r for r in records if r.get("request_id", "") not in bad_ids]
        dropped = before - len(records)
        if dropped:
            LAST_LOAD_DIAGNOSTIC = (
                f"dropped {dropped} conversation(s) with detection_status != 'success'"
            )

    return records


def _load_broken_conversation_ids() -> set[str]:
    """
    Scan all detection_metadata xlsx files for the `detection_status` column
    and return the set of conversation_ids whose status is anything other
    than 'success' (e.g., 'partial', 'error'). Used by load_conversations to
    suppress incomplete-detection conversations from static mode.
    """
    p = config.DETECTION_METADATA_PATH
    candidates: list[Path] = []
    if p.exists():
        if p.is_dir():
            candidates.extend(p.glob("*detection_metadata.xlsx"))
        else:
            candidates.append(p)
            for sib_name in ("threat_detection_metadata.xlsx",
                             "benign_detection_metadata.xlsx"):
                sib = p.with_name(sib_name)
                if sib.exists() and sib != p:
                    candidates.append(sib)
    elif p.parent.exists() and p.parent.is_dir():
        candidates.extend(p.parent.glob("*detection_metadata.xlsx"))

    if not candidates:
        return set()

    from openpyxl import load_workbook  # noqa: PLC0415

    bad_ids: set[str] = set()
    for path in candidates:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            continue
        idx = {name: i for i, name in enumerate(header) if name}
        if "conversation_id" not in idx or "detection_status" not in idx:
            wb.close()
            continue
        cid_i = idx["conversation_id"]
        status_i = idx["detection_status"]
        for row in rows:
            cid = row[cid_i]
            status = row[status_i]
            if cid in (None, ""):
                continue
            if status is None or str(status).strip().lower() != "success":
                bad_ids.add(str(cid))
        wb.close()
    return bad_ids


# ---------------------------------------------------------------------------
# 3. Pre-computed detection results (static mode overlay)
# ---------------------------------------------------------------------------

def load_detection_results(conversation_id: str) -> dict[int, dict[str, Any]] | None:
    """
    Return per-turn detection results for one conversation, drawn from
    detection_metadata.xlsx. The on-disk schema is one row per conversation
    where each per-turn field (prediction / p_detected / latency_ms / ...) is
    a JSON-serialised list of length n_turns, and each (objective, stance)
    pair has its own column for each field.

    Returned shape (consumed by render.render_detection_panel):
        {
            turn_index: {
                "<objective>__<stance>": {
                    "prediction": "Y"|"N"|"E"|"X",
                    "p_detected": float,
                    "p_not_detected": float,
                    "latency_ms": int,
                },
                ...
            },
            ...
        }
    Turns where every detector logged null are omitted from the outer dict.
    Returns None if no detection metadata file exists for either label.
    """
    candidates: list[Path] = []
    p = config.DETECTION_METADATA_PATH
    if p.exists():
        if p.is_dir():
            for name in ("threat_detection_metadata.xlsx",
                         "benign_detection_metadata.xlsx",
                         "detection_metadata.xlsx"):
                f = p / name
                if f.exists():
                    candidates.append(f)
        else:
            candidates.append(p)
            for sib_name in ("threat_detection_metadata.xlsx",
                             "benign_detection_metadata.xlsx"):
                sib = p.with_name(sib_name)
                if sib.exists() and sib != p:
                    candidates.append(sib)
    elif p.parent.exists() and p.parent.is_dir():
        for f in p.parent.glob("*detection_metadata.xlsx"):
            candidates.append(f)

    if not candidates:
        return None

    from openpyxl import load_workbook  # noqa: PLC0415

    objectives = ("policy_violation", "social_engineering")
    stances = ("high_precision", "balanced", "high_recall")
    keys = [f"{o}__{s}" for o in objectives for s in stances]
    fields = ("prediction", "p_detected", "p_not_detected", "latency_ms")

    for path in candidates:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            continue
        idx = {name: i for i, name in enumerate(header) if name}
        if "conversation_id" not in idx:
            wb.close()
            continue
        match = None
        for row in rows:
            if row[idx["conversation_id"]] == conversation_id:
                match = row
                break
        wb.close()
        if not match:
            continue

        n_turns_v = match[idx["n_turns"]] if "n_turns" in idx else None

        def _list(col: str) -> list[Any]:
            if col not in idx or match[idx[col]] in (None, ""):
                return []
            try:
                v = json.loads(match[idx[col]])
                return v if isinstance(v, list) else []
            except (TypeError, json.JSONDecodeError):
                return []

        out: dict[int, dict[str, Any]] = {}
        for key in keys:
            per_field = {f: _list(f"{key}__{f}") for f in fields}
            length = max((len(v) for v in per_field.values()), default=0)
            if n_turns_v is not None:
                try:
                    length = max(length, int(n_turns_v))
                except (TypeError, ValueError):
                    pass
            for i in range(length):
                preds = per_field["prediction"]
                pred = preds[i] if i < len(preds) else None
                if pred in (None, ""):
                    continue
                slot = out.setdefault(i, {})
                slot[key] = {
                    "prediction":     pred,
                    "p_detected":     per_field["p_detected"][i] if i < len(per_field["p_detected"]) else None,
                    "p_not_detected": per_field["p_not_detected"][i] if i < len(per_field["p_not_detected"]) else None,
                    "latency_ms":     per_field["latency_ms"][i] if i < len(per_field["latency_ms"]) else None,
                }
        return out or None

    return None


def index_stance_detections() -> dict[str, set[str]]:
    """
    Read every detection_metadata xlsx file and return a mapping of
        conversation_id -> set of stance keys where at least one turn's
        prediction was "Y" (or "S") for any objective.

    Conversations with no detection rows are simply absent from the result.
    Used by the static-mode UI to filter conversations by stance.
    """
    candidates: list[Path] = []
    p = config.DETECTION_METADATA_PATH
    if p.exists():
        if p.is_dir():
            for f in p.glob("*detection_metadata.xlsx"):
                candidates.append(f)
        else:
            candidates.append(p)
            for sib_name in ("threat_detection_metadata.xlsx",
                             "benign_detection_metadata.xlsx"):
                sib = p.with_name(sib_name)
                if sib.exists() and sib != p:
                    candidates.append(sib)
    elif p.parent.exists() and p.parent.is_dir():
        for f in p.parent.glob("*detection_metadata.xlsx"):
            candidates.append(f)

    if not candidates:
        return {}

    from openpyxl import load_workbook  # noqa: PLC0415

    objectives = ("policy_violation", "social_engineering")

    out: dict[str, set[str]] = {}
    for path in candidates:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            wb.close()
            continue
        idx = {name: i for i, name in enumerate(header) if name}
        if "conversation_id" not in idx:
            wb.close()
            continue

        # Detection columns follow the pattern <obj>__<stance>__prediction.
        pred_columns: list[tuple[str, str, int]] = []  # (obj, stance, col_idx)
        for col_name, col_idx in idx.items():
            if not isinstance(col_name, str):
                continue
            if not col_name.endswith("__prediction"):
                continue
            base = col_name[: -len("__prediction")]
            for obj in objectives:
                if base.startswith(f"{obj}__"):
                    stance = base[len(obj) + 2:]
                    pred_columns.append((obj, stance, col_idx))
                    break

        for row in rows:
            cid = row[idx["conversation_id"]]
            if cid in (None, ""):
                continue
            for _obj, stance, ci in pred_columns:
                cell = row[ci]
                if cell in (None, ""):
                    continue
                try:
                    preds = json.loads(cell)
                except (TypeError, json.JSONDecodeError):
                    continue
                if not isinstance(preds, list):
                    continue
                if any(p in ("Y", "S") for p in preds):
                    out.setdefault(str(cid), set()).add(stance)
        wb.close()

    return out


# ---------------------------------------------------------------------------
# 4. Dynamic generation
# ---------------------------------------------------------------------------

def _materialise_dimensions(dimensions: dict, selection: dict) -> dict:
    """
    Return a shallow copy of `dimensions` with any custom user-entered
    strings injected under a synthetic '__custom__' key in the relevant
    dimension block. The generator's `render_one` looks up
    `dimensions[dim][key]`, so this is what makes "Custom" choices work.
    """
    field_to_dim_threat = {
        "scenario_key":          "scenarios",
        "representative_key":    "representatives",
        "caller_key":            "threat_caller_profile",
        "benign_context_key":    "benign_context_levels",
        "cialdini_emphasis_key": "cialdini_emphasis",
        "turn_count_key":        "turn_counts",
    }
    # Pick the caller dim from the template's declared dim_map.
    tpl_block = dimensions["prompt_templates"][selection["prompt_template_key"]]
    tpl_dims = tpl_block.get("dimensions", {}) if isinstance(tpl_block, dict) else {}
    caller_dim = tpl_dims.get("caller_key")
    field_to_dim = dict(field_to_dim_threat)
    if caller_dim:
        field_to_dim["caller_key"] = caller_dim

    out: dict[str, Any] = {k: v for k, v in dimensions.items()}
    for field, dim in field_to_dim.items():
        if selection.get(field) != "__custom__":
            continue
        custom_text = selection.get(field.replace("_key", "_custom"), "")
        if not custom_text:
            continue
        out[dim] = dict(out.get(dim, {}))
        out[dim]["__custom__"] = custom_text
    return out


def generate_conversation(selection: dict[str, Any]) -> dict[str, Any]:
    """
    Dynamic mode: generate a single conversation matching the existing
    Conversation Generation module's behaviour.

    `selection` keys consumed:
      - prompt_template_key, scenario_key, representative_key, caller_key,
        benign_context_key, cialdini_emphasis_key, turn_count_key
        (any subset; only those declared by the template are used)
      - *_custom: custom text (when the corresponding *_key == "__custom__")
      - flavor, model, temperature, top_p

    Returns a normalised record matching the load_conversations() schema.
    """
    cg = _cg()
    dims = load_dimensions()
    dims_for_call = _materialise_dimensions(dims, selection)

    template_key = selection["prompt_template_key"]
    tpl_block = dims["prompt_templates"][template_key]
    declared = tpl_block.get("dimensions", {}) if isinstance(tpl_block, dict) else {}

    # Fill in safe defaults for any dimension the template doesn't declare.
    def _key_or_default(field: str, dim_name: str) -> str:
        if field in declared:
            return selection.get(field, "")
        block = dims_for_call.get(dim_name, {})
        return next(iter(block.keys())) if block else ""

    system_prompt, user_prompt = cg.render_one(
        dims_for_call,
        template_key=template_key,
        scenario_key=_key_or_default("scenario_key", "scenarios"),
        representative_key=_key_or_default("representative_key", "representatives"),
        caller_key=selection.get("caller_key", ""),
        benign_context_key=_key_or_default("benign_context_key", "benign_context_levels"),
        cialdini_emphasis_key=_key_or_default("cialdini_emphasis_key", "cialdini_emphasis"),
        turn_count_key=_key_or_default("turn_count_key", "turn_counts"),
        flavor=selection.get("flavor", ""),
    )

    turn_count_key = selection.get("turn_count_key") or _key_or_default(
        "turn_count_key", "turn_counts"
    )
    turn_count = int(dims_for_call["turn_counts"][turn_count_key])

    parsed, usage = cg.call_openai(
        system_prompt=system_prompt,
        user_prompt=user_prompt,
        model=selection["model"],
        temperature=float(selection["temperature"]),
        top_p=float(selection["top_p"]),
        turn_count=turn_count,
    )

    request_id = f"dynamic-{uuid.uuid4().hex[:12]}"
    record = {
        "request_id": request_id,
        "prompt_template_key": template_key,
        "selection": {k: selection.get(k, "") for k in (
            "scenario_key", "representative_key", "caller_key",
            "benign_context_key", "cialdini_emphasis_key", "turn_count_key",
        ) if selection.get(k)},
        "flavor": selection.get("flavor", ""),
        "model": selection["model"],
        "temperature": selection["temperature"],
        "top_p": selection["top_p"],
        "system_prompt": system_prompt,
        "user_prompt": user_prompt,
        "conversation": parsed,
        "usage": usage,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "_source_label": (
            "threat" if "threat" in template_key.lower() else
            ("benign" if "benign" in template_key.lower() else "dynamic")
        ),
    }
    return _normalise_record(record)


# ---------------------------------------------------------------------------
# 5. Live detection (dynamic mode)
# ---------------------------------------------------------------------------

# Cached detection-pipeline state.
_DET_CACHE: dict[str, Any] = {}


def _load_detection_dims() -> tuple[Any, Any, Any]:
    """Return (client, policy_dims, se_dims), cached after first call."""
    if "policy" in _DET_CACHE:
        return _DET_CACHE["client"], _DET_CACHE["policy"], _DET_CACHE["se"]

    det = _det()
    det_dir = _PROJECT_ROOT / "detection_pipeline"

    def _find(*candidates: str) -> Path:
        for name in candidates:
            f = det_dir / name
            if f.exists():
                return f
        # Fall back: scan for any json whose name contains a candidate stem.
        for f in det_dir.glob("*.json"):
            for c in candidates:
                stem = c.replace(".json", "")
                if stem in f.name:
                    return f
        raise FileNotFoundError(
            f"Could not find detection dimensions in {det_dir}. "
            f"Looked for: {candidates}."
        )

    policy_path = _find(
        "policy_detection_prompt_dimensions.json",
        "policy_violation_dimensions.json",
        "policy_dimensions.json",
        "detection_policy.json",
    )
    se_path = _find(
        "se_detection_prompt_dimensions.json",
        "social_engineering_dimensions.json",
        "se_dimensions.json",
        "detection_se.json",
    )

    from openai import OpenAI  # noqa: PLC0415
    client = OpenAI()
    policy_dims = det.load_detection_dimensions(policy_path)
    se_dims = det.load_detection_dimensions(se_path)

    _DET_CACHE["client"] = client
    _DET_CACHE["policy"] = policy_dims
    _DET_CACHE["se"] = se_dims
    return client, policy_dims, se_dims


def _conv_dict_through_turn(record: dict, k: int) -> dict:
    """
    Return a `{turn_0: ..., turn_1: ..., ..., turn_k: ...}` dict copied from
    the original turn dicts (without the speaker/text we added in
    normalisation). Feeds the detection module, which expects the raw
    turn_N schema.
    """
    conv = record["conversation"]
    out: dict[str, Any] = {}
    for i in range(k + 1):
        key = f"turn_{i}"
        if key in conv:
            t = dict(conv[key])
        else:
            # Reconstruct from the normalised list if the original keys are
            # no longer there (e.g., dynamic-mode in-memory records).
            t = dict(conv["turns"][i])
        t.pop("speaker", None)
        t.pop("text", None)
        out[key] = t
    return out


def _materialise_stance(dims: dict, custom_stance_text: str | None) -> dict:
    """
    Return a copy of dims with a "__custom__" stance injected if
    `custom_stance_text` is non-empty. Mirrors the pattern used for custom
    dimension values in generation.
    """
    if not custom_stance_text:
        return dims
    out = dict(dims)
    stances_block = dict(out.get("stances", {}))
    stances_block["__custom__"] = {"instruction": custom_stance_text}
    out["stances"] = stances_block
    return out


def list_detection_stances() -> list[str]:
    """
    Return the union of preset stance keys declared in the policy_violation
    and social_engineering detection dimension files. Used by the UI to
    populate the stance dropdown. Does not instantiate an OpenAI client.
    """
    det_dir = _PROJECT_ROOT / "detection_pipeline"
    fnames = (
        ("policy_detection_prompt_dimensions.json",
         "policy_violation_dimensions.json", "policy_dimensions.json"),
        ("se_detection_prompt_dimensions.json",
         "social_engineering_dimensions.json", "se_dimensions.json"),
    )
    keys: list[str] = []
    for candidates in fnames:
        path = None
        for n in candidates:
            f = det_dir / n
            if f.exists():
                path = f
                break
        if not path:
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                d = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        for k in (d.get("stances", {}) or {}).keys():
            if k not in keys:
                keys.append(k)
    if not keys:
        keys = ["high_precision", "balanced", "high_recall"]
    return keys


def score_turns_through(
    conversation: dict[str, Any],
    target_turn_index: int,
    *,
    stances: list[str] | None = None,
    custom_stance_text: str = "",
    model: str | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Run the chosen stance variants of the appropriate detector on the
    conversation truncated through `target_turn_index`, and return a dict
    shaped like one entry of `load_detection_results`'s output.

    Args:
        conversation: the conversation record (must have its `turn_N` keys
            intact — `_conv_dict_through_turn` reconstructs from the
            normalised list as a fallback).
        target_turn_index: which turn to score through (inclusive).
        stances: list of stance keys to run. None or empty means run all
            three preset stances ("high_precision", "balanced", "high_recall").
            Pass ["__custom__"] to run only the custom stance (requires
            `custom_stance_text` non-empty).
        custom_stance_text: free-text stance instruction. When non-empty
            it's injected into the loaded detection dims under
            `stances["__custom__"]`.
        model: detection model override; defaults to
            config.DEFAULT_DETECTION_MODEL.

    For turn 0 (rep greeting) returns {} — there's nothing to score.
    For rep turns runs only `policy_violation__*`; for caller turns runs
    only `social_engineering__*`.
    """
    if target_turn_index <= 0:
        return {}

    speaker = _speaker_for_turn(target_turn_index)
    if speaker == "representative":
        objective = "policy_violation"
    elif speaker == "caller":
        objective = "social_engineering"
    else:
        return {}

    client, policy_dims, se_dims = _load_detection_dims()
    det = _det()
    base_dims = policy_dims if objective == "policy_violation" else se_dims
    dims_for_obj = _materialise_stance(base_dims, custom_stance_text)

    if not stances:
        stances = ["high_precision", "balanced", "high_recall"]

    # If "__custom__" is in the requested list but no custom text supplied,
    # silently drop it.
    if "__custom__" in stances and not custom_stance_text:
        stances = [s for s in stances if s != "__custom__"]

    if not stances:
        return {}

    detection_model = model or config.DEFAULT_DETECTION_MODEL

    conv_dict = _conv_dict_through_turn(conversation, target_turn_index)
    turn_keys = [f"turn_{i}" for i in range(target_turn_index + 1)]
    user_prompt = det.transcript_through_turn(conv_dict, turn_keys, target_turn_index)

    out: dict[str, dict[str, Any]] = {}
    for stance in stances:
        # Skip stances that aren't actually present in the dims (the user
        # might have asked for a preset that this objective doesn't define).
        if stance not in (dims_for_obj.get("stances", {}) or {}):
            continue
        system_prompt = det.assemble_system_prompt(dims_for_obj, stance)
        res, _attempts, err = det.call_detector_with_retries(
            client=client,
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            model=detection_model,
            max_attempts=2,
            backoff=(2,),
            top_logprobs=10,
        )
        key = f"{objective}__{stance}"
        if res is None:
            out[key] = {
                "prediction":     "X",
                "p_detected":     None,
                "p_not_detected": None,
                "latency_ms":     None,
            }
            continue
        out[key] = {
            "prediction":     res.get("prediction"),
            "p_detected":     res.get("p_detected"),
            "p_not_detected": res.get("p_not_detected"),
            "latency_ms":     res.get("latency_ms"),
        }
    return out
