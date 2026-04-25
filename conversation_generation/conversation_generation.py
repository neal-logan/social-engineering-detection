"""
Conversation generation pipeline (stage 1 of the broader social-engineering
detection project).

Designed for separate threat and benign runs: each run is configured with
its own template key, output paths, flavor-sampling strategy, and flavor
count. The same module powers both.

Schema enforcement:
    Each call uses OpenAI's structured outputs against a Pydantic model
    that is built dynamically from the request's turn count. The model has
    one field per turn position (turn_0 ... turn_{N-1}) with a fixed shape
    by parity:
      - turn_0: representative greeting (passage only)
      - odd turns: caller turn (passage + 7 cialdini flags)
      - even turns >= 2: representative turn (passage + 3 improper flags)

Retries:
    Each request is attempted up to MAX_ATTEMPTS times. Any exception
    (network error, schema validation failure, refusal, etc.) triggers a
    retry with exponential backoff (2, 4, 8, 16, 32, then 60s for the
    rest). Final failure is recorded in the metadata tracker.

Concurrency and durability:
    The generation loop runs requests in batches; within a batch, requests
    execute concurrently in a thread pool. Results from the OpenAI SDK
    are buffered in memory and flushed to disk every FLUSH_EVERY completed
    requests (by default), reducing per-request disk writes by ~20x. Both
    files are written atomically (temp file + os.replace) so an interrupt
    cannot leave them half-written.
"""

from __future__ import annotations

import itertools
import json
import os
import random
import shutil
import tempfile
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field, create_model
from tqdm.auto import tqdm


CIALDINI_FLAGS = [
    "cialdini_reciprocity",
    "cialdini_commitment_consistency",
    "cialdini_social_proof",
    "cialdini_authority",
    "cialdini_liking",
    "cialdini_scarcity",
    "cialdini_unity",
]

IMPROPER_FLAGS = [
    "improper_authentication",
    "improper_disclosure",
    "improper_action",
]


def _flag_metric_columns() -> list[str]:
    """Returns the 20 metric columns: 10 _sum and 10 _by_turn."""
    cols: list[str] = []
    for flag in CIALDINI_FLAGS + IMPROPER_FLAGS:
        cols.append(f"{flag}_sum")
        cols.append(f"{flag}_by_turn")
    return cols


METADATA_COLUMNS = [
    "request_id",
    "replicate_index",
    "prompt_template_key",
    "scenario",
    "representative",
    "caller",
    "benign_context",
    "cialdini_emphasis",
    "turn_count_value",
    "flavor",
    "model",
    "temperature",
    "top_p",
    "status",            # pending | success | error
    "attempts",
    "last_error",
    "generated_at_utc",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    *_flag_metric_columns(),
    "system_prompt",
    "user_prompt",
]

_TEMPLATE_DIMENSION_FIELDS = [
    "scenario_key", "representative_key", "caller_key",
    "benign_context_key", "cialdini_emphasis_key", "turn_count_key",
]

_FIELD_TO_METADATA = {
    "scenario_key":          "scenario",
    "representative_key":    "representative",
    "caller_key":            "caller",
    "benign_context_key":    "benign_context",
    "cialdini_emphasis_key": "cialdini_emphasis",
    "turn_count_key":        "turn_count_value",
}

# Flavor sampling strategies
FLAVOR_DETERMINISTIC = "deterministic"
FLAVOR_RESAMPLED     = "resampled"

# Retry behavior
MAX_ATTEMPTS = 10
_BACKOFF_SCHEDULE = [2, 4, 8, 16, 32]  # seconds; further attempts use 60s

# Concurrency and durability defaults
BATCH_SIZE   = 5    # concurrent OpenAI calls per batch
FLUSH_EVERY  = 20   # flush metadata + conversations to disk every N completed requests


# ---------------------------------------------------------------------------
# Pydantic models for structured output (built dynamically per turn count)
# ---------------------------------------------------------------------------

class _RepGreetingTurn(BaseModel):
    """Representative greeting turn (turn 0)."""
    passage: str

    model_config = {"extra": "forbid"}


class _CallerTurn(BaseModel):
    """Caller turn: passage + 7 cialdini flags."""
    passage: str
    cialdini_reciprocity: bool
    cialdini_commitment_consistency: bool
    cialdini_social_proof: bool
    cialdini_authority: bool
    cialdini_liking: bool
    cialdini_scarcity: bool
    cialdini_unity: bool

    model_config = {"extra": "forbid"}


class _RepResponseTurn(BaseModel):
    """Representative response turn: passage + 3 improper flags."""
    passage: str
    improper_authentication: bool
    improper_disclosure: bool
    improper_action: bool

    model_config = {"extra": "forbid"}


@lru_cache(maxsize=16)
def conversation_model_for(turn_count: int):
    """
    Build (and cache) a Pydantic model with one required field per turn
    position. Field shape varies by parity:
      - turn_0: _RepGreetingTurn
      - odd turns (1, 3, 5, ...): _CallerTurn
      - even turns >= 2: _RepResponseTurn
    """
    if turn_count < 1:
        raise ValueError(f"turn_count must be >= 1, got {turn_count}")

    fields: dict[str, Any] = {}
    for i in range(turn_count):
        if i == 0:
            cls = _RepGreetingTurn
        elif i % 2 == 1:
            cls = _CallerTurn
        else:
            cls = _RepResponseTurn
        fields[f"turn_{i}"] = (cls, Field(...))

    Model = create_model(
        f"Conversation{turn_count}Turn",
        __config__={"extra": "forbid"},
        **fields,
    )
    return Model


# ---------------------------------------------------------------------------
# Loading inputs
# ---------------------------------------------------------------------------

def load_dimensions(path: str | Path) -> dict[str, Any]:
    """Load the prompt-dimensions JSON file."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _template_text(dimensions: dict, template_key: str) -> str:
    return dimensions["prompt_templates"][template_key]["template"]


def _template_dim_map(dimensions: dict, template_key: str) -> dict[str, str]:
    return dimensions["prompt_templates"][template_key]["dimensions"]


def render_system_prompt(template: str, replacements: dict[str, str]) -> str:
    out = template
    for placeholder, value in replacements.items():
        out = out.replace("{" + placeholder + "}", str(value))
    return out


def build_user_prompt(flavor: str) -> str:
    return f"flavor: {flavor}"


# ---------------------------------------------------------------------------
# Flavor sampling
# ---------------------------------------------------------------------------

def _sample_flavors(
    flavor_pool: list[str],
    n: int,
    rng: random.Random,
) -> list[str]:
    if n > len(flavor_pool):
        raise ValueError(
            f"Requested {n} flavors but only {len(flavor_pool)} available."
        )
    return rng.sample(flavor_pool, n)


# ---------------------------------------------------------------------------
# Enumeration
# ---------------------------------------------------------------------------

def enumerate_requests(
    dimensions: dict,
    template_key: str,
    flavor_count: int,
    flavor_strategy: str,
    model: str,
    temperature: float,
    top_p: float,
    replicates: int = 1,
    flavor_seed: int | None = None,
) -> list[dict]:
    if flavor_strategy not in (FLAVOR_DETERMINISTIC, FLAVOR_RESAMPLED):
        raise ValueError(f"Unknown flavor_strategy: {flavor_strategy}")

    dim_map = _template_dim_map(dimensions, template_key)
    flavor_pool: list[str] = dimensions["flavors"]
    rng = random.Random(flavor_seed)

    if flavor_strategy == FLAVOR_DETERMINISTIC:
        fixed_flavors = _sample_flavors(flavor_pool, flavor_count, rng)
    else:
        fixed_flavors = None

    axes: list[tuple[str, list[str]]] = []
    for field in _TEMPLATE_DIMENSION_FIELDS:
        if field not in dim_map:
            continue
        dim_name = dim_map[field]
        axes.append((field, list(dimensions[dim_name].keys())))
    names = [a[0] for a in axes]
    value_lists = [a[1] for a in axes]

    requests: list[dict] = []
    for combo in itertools.product(*value_lists):
        selection_keys = dict(zip(names, combo))
        if flavor_strategy == FLAVOR_DETERMINISTIC:
            combo_flavors = fixed_flavors
        else:
            combo_flavors = _sample_flavors(flavor_pool, flavor_count, rng)

        for flavor in combo_flavors:
            for r in range(replicates):
                row = {
                    "request_id": str(uuid.uuid4()),
                    "replicate_index": r,
                    "prompt_template_key": template_key,
                    "scenario": "",
                    "representative": "",
                    "caller": "",
                    "benign_context": "",
                    "cialdini_emphasis": "",
                    "turn_count_value": "",
                    "flavor": flavor,
                    "model": model,
                    "temperature": temperature,
                    "top_p": top_p,
                    "status": "pending",
                    "attempts": 0,
                    "last_error": "",
                    "generated_at_utc": "",
                    "system_prompt": "",
                    "user_prompt": "",
                }
                for field, key in selection_keys.items():
                    dim_name = dim_map[field]
                    metadata_col = _FIELD_TO_METADATA[field]
                    row[metadata_col] = dimensions[dim_name][key]
                requests.append(row)

    return requests


def attach_prompts(requests: list[dict], dimensions: dict) -> None:
    for req in requests:
        tkey = req["prompt_template_key"]
        template = _template_text(dimensions, tkey)
        turn_count = int(req["turn_count_value"])
        replacements = {
            "scenario":          req["scenario"],
            "representative":    req["representative"],
            "caller":            req["caller"],
            "benign_context":    req["benign_context"],
            "cialdini_emphasis": req["cialdini_emphasis"],
            "turn_count":        str(turn_count),
            "last_turn_index":   str(turn_count - 1),
        }
        req["system_prompt"] = render_system_prompt(template, replacements)
        req["user_prompt"] = build_user_prompt(req["flavor"])


# ---------------------------------------------------------------------------
# One-off rendering (for the dev notebook)
# ---------------------------------------------------------------------------

def render_one(
    dimensions: dict,
    template_key: str,
    *,
    scenario_key: str,
    representative_key: str,
    caller_key: str,
    benign_context_key: str,
    cialdini_emphasis_key: str,
    turn_count_key: str,
    flavor: str,
) -> tuple[str, str]:
    dim_map = _template_dim_map(dimensions, template_key)
    template = _template_text(dimensions, template_key)
    selection = {
        "scenario_key": scenario_key,
        "representative_key": representative_key,
        "caller_key": caller_key,
        "benign_context_key": benign_context_key,
        "cialdini_emphasis_key": cialdini_emphasis_key,
        "turn_count_key": turn_count_key,
    }
    placeholder_for_field = {
        "scenario_key": "scenario",
        "representative_key": "representative",
        "caller_key": "caller",
        "benign_context_key": "benign_context",
        "cialdini_emphasis_key": "cialdini_emphasis",
        "turn_count_key": "turn_count",
    }
    replacements: dict[str, str] = {}
    turn_count_value: int | None = None
    for field, key in selection.items():
        if field not in dim_map:
            continue
        dim_name = dim_map[field]
        value = dimensions[dim_name][key]
        replacements[placeholder_for_field[field]] = str(value)
        if field == "turn_count_key":
            turn_count_value = int(value)
    if turn_count_value is not None:
        replacements["last_turn_index"] = str(turn_count_value - 1)
        replacements["turn_count"] = str(turn_count_value)
    return (
        render_system_prompt(template, replacements),
        build_user_prompt(flavor),
    )


# ---------------------------------------------------------------------------
# Metadata (.xlsx) tracker
# ---------------------------------------------------------------------------

def _atomic_save_workbook(wb: Workbook, path: Path) -> None:
    """
    Save an openpyxl Workbook to `path` atomically: write to a sibling
    temp file, then os.replace into place. An interrupt during the save
    leaves either the previous valid file or the new one - never half-written.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        prefix=path.name + ".",
        suffix=".tmp",
        dir=str(path.parent),
    )
    os.close(fd)
    tmp_path = Path(tmp_name)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        try:
            tmp_path.unlink()
        except FileNotFoundError:
            pass
        raise


def _atomic_write_text(text: str, path: Path) -> None:
    """Atomically write a string to a file via a sibling temp file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(
        prefix=path.name + ".",
        suffix=".tmp",
        dir=str(path.parent),
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(text)
        os.replace(tmp_name, path)
    except Exception:
        try:
            os.unlink(tmp_name)
        except FileNotFoundError:
            pass
        raise


def write_metadata_xlsx(requests: list[dict], path: str | Path) -> None:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "requests"
    ws.append(METADATA_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for req in requests:
        ws.append([req.get(col, "") for col in METADATA_COLUMNS])
    ws.freeze_panes = "A2"

    default_widths = {
        "request_id": 38, "replicate_index": 8, "prompt_template_key": 14,
        "scenario": 60, "representative": 60, "caller": 60,
        "benign_context": 60, "cialdini_emphasis": 60,
        "turn_count_value": 8, "flavor": 28,
        "model": 14, "temperature": 12, "top_p": 8, "status": 10,
        "attempts": 10, "last_error": 40, "generated_at_utc": 26,
        "input_tokens": 12, "output_tokens": 12, "total_tokens": 12,
        "system_prompt": 80, "user_prompt": 40,
    }
    # Sum columns are narrow ints; by_turn columns are short comma-separated lists.
    for flag in CIALDINI_FLAGS + IMPROPER_FLAGS:
        default_widths[f"{flag}_sum"] = 6
        default_widths[f"{flag}_by_turn"] = 32
    for i, col in enumerate(METADATA_COLUMNS, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = default_widths.get(col, 16)

    wrap_cols = {
        "scenario", "representative", "caller", "benign_context",
        "cialdini_emphasis", "system_prompt", "user_prompt", "last_error",
    }
    for col_idx, col_name in enumerate(METADATA_COLUMNS, start=1):
        if col_name in wrap_cols:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    _atomic_save_workbook(wb, path)


def read_metadata_xlsx(path: str | Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb["requests"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out = []
    for row in rows[1:]:
        out.append({header[i]: (row[i] if row[i] is not None else "") for i in range(len(header))})
    return out


def update_metadata_row(path: str | Path, request_id: str, updates: dict) -> None:
    """Update a single row. Atomic: saves via temp file + os.replace."""
    update_metadata_rows(path, {request_id: updates})


def update_metadata_rows(path: str | Path, updates_by_id: dict[str, dict]) -> None:
    """
    Apply updates to many rows in one workbook open/save cycle. Atomic.
    `updates_by_id` maps request_id -> {column_name: new_value}. Rows whose
    request_id isn't in the workbook are silently skipped.
    """
    if not updates_by_id:
        return
    path = Path(path)
    wb = load_workbook(path)
    ws = wb["requests"]
    header = [c.value for c in ws[1]]
    id_col = header.index("request_id") + 1
    col_index = {name: idx + 1 for idx, name in enumerate(header)}

    remaining = dict(updates_by_id)
    for row_idx in range(2, ws.max_row + 1):
        if not remaining:
            break
        rid = ws.cell(row=row_idx, column=id_col).value
        upd = remaining.pop(rid, None)
        if upd is None:
            continue
        for key, value in upd.items():
            ci = col_index.get(key)
            if ci is not None:
                ws.cell(row=row_idx, column=ci).value = value
    _atomic_save_workbook(wb, path)


def ensure_metadata_columns(path: str | Path) -> list[str]:
    """
    Make sure the workbook at `path` has every column in the current
    METADATA_COLUMNS schema. Any missing columns are appended at the end of
    the header row (existing data is preserved). Returns the list of column
    names that were added.
    """
    path = Path(path)
    wb = load_workbook(path)
    ws = wb["requests"]
    header = [c.value for c in ws[1]]
    missing = [c for c in METADATA_COLUMNS if c not in header]
    if not missing:
        return []
    # Append missing columns at the end and bold their headers.
    for col_name in missing:
        new_col_idx = ws.max_column + 1
        cell = ws.cell(row=1, column=new_col_idx, value=col_name)
        cell.font = Font(bold=True)
        default_widths = {
            "input_tokens": 12, "output_tokens": 12, "total_tokens": 12,
        }
        for flag in CIALDINI_FLAGS + IMPROPER_FLAGS:
            default_widths[f"{flag}_sum"] = 6
            default_widths[f"{flag}_by_turn"] = 32
        if col_name in default_widths:
            ws.column_dimensions[cell.column_letter].width = default_widths[col_name]
    _atomic_save_workbook(wb, path)
    return missing


# ---------------------------------------------------------------------------
# Conversation store (.json)
# ---------------------------------------------------------------------------

def init_conversation_store(path: str | Path) -> None:
    """Create the conversations JSON file (and parent dir) if missing."""
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.exists():
        _atomic_write_text(json.dumps({"conversations": {}}, indent=2), p)


def load_conversation_store(path: str | Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_conversation(path: str | Path, request_id: str, record: dict) -> None:
    """Append / overwrite one conversation. Atomic. Convenience wrapper."""
    save_conversations(path, {request_id: record})


def save_conversations(path: str | Path, records: dict[str, dict]) -> None:
    """
    Append / overwrite many conversations in one read/write cycle. Atomic.
    `records` maps request_id -> conversation record.
    """
    if not records:
        return
    path = Path(path)
    store = load_conversation_store(path)
    store["conversations"].update(records)
    _atomic_write_text(
        json.dumps(store, indent=2, ensure_ascii=False),
        path,
    )


# ---------------------------------------------------------------------------
# Flag extraction
# ---------------------------------------------------------------------------

def extract_flag_metrics(conversation: dict, turn_count: int) -> dict[str, Any]:
    """
    Walk a parsed conversation dict (turn_0, turn_1, ...) and produce
    metric values for every cialdini and improper flag.

    Returns a dict with:
      - "{flag}_sum":     int count of turns where the flag was True
                          (only counted on turns where the flag applies)
      - "{flag}_by_turn": comma-separated string of length turn_count,
                          with "0" or "1" per position. Position i is "1"
                          iff the flag is applicable to turn i and is True
                          on that turn; "0" otherwise (including turns
                          where the flag does not apply).

    Cialdini flags are applicable on caller turns (odd indices >= 1).
    Improper flags are applicable on rep response turns (even indices >= 2).
    Turn 0 is the rep greeting and contributes "0" to every flag.

    Missing turns or missing flag fields are treated as not-True (no
    contribution to the sum, "0" in the by-turn string). This makes the
    function robust to partial or malformed conversations.
    """
    sums: dict[str, int] = {f: 0 for f in CIALDINI_FLAGS + IMPROPER_FLAGS}
    by_turn: dict[str, list[str]] = {f: ["0"] * turn_count for f in CIALDINI_FLAGS + IMPROPER_FLAGS}

    for i in range(turn_count):
        turn = conversation.get(f"turn_{i}")
        if not isinstance(turn, dict):
            continue
        if i == 0:
            continue  # greeting; nothing to extract
        if i % 2 == 1:
            # caller turn: cialdini flags apply
            for flag in CIALDINI_FLAGS:
                if turn.get(flag) is True:
                    sums[flag] += 1
                    by_turn[flag][i] = "1"
        else:
            # rep response turn: improper flags apply
            for flag in IMPROPER_FLAGS:
                if turn.get(flag) is True:
                    sums[flag] += 1
                    by_turn[flag][i] = "1"

    metrics: dict[str, Any] = {}
    for flag in CIALDINI_FLAGS + IMPROPER_FLAGS:
        metrics[f"{flag}_sum"] = sums[flag]
        metrics[f"{flag}_by_turn"] = ",".join(by_turn[flag])
    return metrics


# ---------------------------------------------------------------------------
# OpenAI call (structured outputs)
# ---------------------------------------------------------------------------

def _get_openai_client():
    from dotenv import load_dotenv
    from openai import OpenAI
    load_dotenv()
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY not set in environment or .env file.")
    return OpenAI(api_key=api_key)


def call_openai(
    system_prompt: str,
    user_prompt: str,
    model: str,
    temperature: float,
    top_p: float,
    turn_count: int,
) -> tuple[dict, dict]:
    """
    One non-streaming structured-output call. Returns (parsed_conversation,
    usage), where parsed_conversation is the validated dict and usage is
    {"input_tokens", "output_tokens", "total_tokens"}.

    Raises on refusal, validation failure, or API error.
    """
    client = _get_openai_client()
    Model = conversation_model_for(turn_count)

    kwargs = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_prompt},
        ],
        "response_format": Model,
    }
    # GPT-5 family currently only supports temperature=1 and top_p=1.
    if temperature != 1:
        kwargs["temperature"] = temperature
    if top_p != 1:
        kwargs["top_p"] = top_p

    completion = client.chat.completions.parse(**kwargs)
    msg = completion.choices[0].message

    if getattr(msg, "refusal", None):
        raise RuntimeError(f"Model refused: {msg.refusal}")
    if msg.parsed is None:
        raise RuntimeError("Structured-output parse returned None.")

    usage_obj = getattr(completion, "usage", None)
    usage = {
        "input_tokens":  getattr(usage_obj, "prompt_tokens", 0) or 0,
        "output_tokens": getattr(usage_obj, "completion_tokens", 0) or 0,
        "total_tokens":  getattr(usage_obj, "total_tokens", 0) or 0,
    }

    return msg.parsed.model_dump(), usage


def _backoff_seconds(attempt: int) -> float:
    """attempt is 1-indexed and refers to the attempt that just FAILED."""
    if attempt - 1 < len(_BACKOFF_SCHEDULE):
        return _BACKOFF_SCHEDULE[attempt - 1]
    return 60.0


def call_openai_with_retries(
    system_prompt: str,
    user_prompt: str,
    model: str,
    temperature: float,
    top_p: float,
    turn_count: int,
    max_attempts: int = MAX_ATTEMPTS,
) -> tuple[dict, dict, int, str]:
    """
    Retry wrapper. Returns (parsed_dict, usage, attempts_used, last_error).
    Only the successful call's usage is recorded (not summed across retries).
    Raises the final exception if all attempts fail.
    """
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            parsed, usage = call_openai(
                system_prompt=system_prompt,
                user_prompt=user_prompt,
                model=model,
                temperature=temperature,
                top_p=top_p,
                turn_count=turn_count,
            )
            return parsed, usage, attempt, ""
        except Exception as e:  # retry on any exception
            last_exc = e
            if attempt < max_attempts:
                time.sleep(_backoff_seconds(attempt))
    assert last_exc is not None
    raise last_exc


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def _process_one_row(row: dict, max_attempts: int) -> dict:
    """
    Worker for a single metadata row. Returns a dict describing the outcome
    so the orchestrator can buffer it without doing any I/O here:

    Success:
        {"request_id", "kind": "success", "metadata_updates", "record"}

    Failure (after retries):
        {"request_id", "kind": "error", "metadata_updates"}
    """
    request_id = row["request_id"]
    prior_attempts = int(row.get("attempts") or 0)
    turn_count = int(row["turn_count_value"])

    try:
        parsed, usage, attempts_used, _ = call_openai_with_retries(
            system_prompt=row["system_prompt"],
            user_prompt=row["user_prompt"],
            model=row["model"],
            temperature=float(row["temperature"]),
            top_p=float(row["top_p"]),
            turn_count=turn_count,
            max_attempts=max_attempts,
        )
    except Exception as e:
        return {
            "request_id": request_id,
            "kind": "error",
            "metadata_updates": {
                "status": "error",
                "attempts": prior_attempts + max_attempts,
                "last_error": f"{type(e).__name__}: {e}",
            },
        }

    record = {
        "request_id": request_id,
        "selection": {k: row[k] for k in [
            "prompt_template_key", "scenario", "representative",
            "caller", "benign_context", "cialdini_emphasis",
            "turn_count_value",
        ]},
        "replicate_index": row["replicate_index"],
        "flavor": row["flavor"],
        "model": row["model"],
        "temperature": row["temperature"],
        "top_p": row["top_p"],
        "system_prompt": row["system_prompt"],
        "user_prompt": row["user_prompt"],
        "conversation": parsed,
        "usage": usage,
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
    }
    metrics = extract_flag_metrics(parsed, turn_count)
    return {
        "request_id": request_id,
        "kind": "success",
        "record": record,
        "metadata_updates": {
            "status": "success",
            "attempts": prior_attempts + attempts_used,
            "last_error": "",
            "generated_at_utc": record["generated_at_utc"],
            "input_tokens":  usage["input_tokens"],
            "output_tokens": usage["output_tokens"],
            "total_tokens":  usage["total_tokens"],
            **metrics,
        },
    }


def _flush_buffers(
    metadata_path: str | Path,
    conversations_path: str | Path,
    pending_metadata: dict[str, dict],
    pending_conversations: dict[str, dict],
) -> None:
    """Flush both pending buffers to disk atomically; clear them in place."""
    if pending_conversations:
        save_conversations(conversations_path, pending_conversations)
        pending_conversations.clear()
    if pending_metadata:
        update_metadata_rows(metadata_path, pending_metadata)
        pending_metadata.clear()


def run_generation_loop(
    metadata_path: str | Path,
    conversations_path: str | Path,
    max_requests: int | None = None,
    max_attempts: int = MAX_ATTEMPTS,
    batch_size: int = BATCH_SIZE,
    flush_every: int = FLUSH_EVERY,
) -> None:
    """
    Walk the metadata tracker, generating any pending or errored rows.

    Concurrency:
        Requests are processed in batches of `batch_size`. Within a batch
        they run concurrently in a thread pool; each call still has its
        own retry loop, so one slow/failing request doesn't block others.
        Batches do not overlap.

    Durability:
        Successful conversations and metadata updates are buffered in
        memory and flushed to disk after each batch completes, but only
        once accumulated work reaches `flush_every` requests (so the
        actual flush cadence is `flush_every` rounded up to the next
        batch boundary). Errored rows are also buffered and flushed at
        the same boundary - they will be retried on the next run.
        Both files are written atomically (temp file + os.replace), so an
        interrupt cannot leave them half-written.
    """
    if batch_size < 1:
        raise ValueError("batch_size must be >= 1")
    if flush_every < 1:
        raise ValueError("flush_every must be >= 1")

    init_conversation_store(conversations_path)
    rows = read_metadata_xlsx(metadata_path)
    pending = [r for r in rows if r["status"] != "success"]
    if max_requests is not None:
        pending = pending[:max_requests]

    pending_metadata: dict[str, dict] = {}
    pending_conversations: dict[str, dict] = {}
    since_flush = 0

    pbar = tqdm(total=len(pending), desc="Generating conversations", unit="conv")
    try:
        for batch_start in range(0, len(pending), batch_size):
            batch = pending[batch_start:batch_start + batch_size]
            with ThreadPoolExecutor(max_workers=batch_size) as ex:
                futures = [ex.submit(_process_one_row, row, max_attempts) for row in batch]
                for fut in as_completed(futures):
                    result = fut.result()
                    pending_metadata[result["request_id"]] = result["metadata_updates"]
                    if result["kind"] == "success":
                        pending_conversations[result["request_id"]] = result["record"]
                    pbar.update(1)
                    since_flush += 1
            if since_flush >= flush_every:
                _flush_buffers(metadata_path, conversations_path,
                               pending_metadata, pending_conversations)
                since_flush = 0
        # Final flush.
        _flush_buffers(metadata_path, conversations_path,
                       pending_metadata, pending_conversations)
    finally:
        # On any unexpected exception (and on KeyboardInterrupt), persist what
        # we have before unwinding, so the next resume picks up cleanly.
        _flush_buffers(metadata_path, conversations_path,
                       pending_metadata, pending_conversations)
        pbar.close()
